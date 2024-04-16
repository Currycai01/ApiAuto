import json

import openpyxl
import pytest
from faker import Faker
from faker import Factory
from faker.providers import BaseProvider

def read_excel(filename,sheetname):
    wb = openpyxl.load_workbook(filename)#加载工作表
    sheet = wb[sheetname]#获取表单
    max_raw = sheet.max_row#获取最大行数

    case_list = []
    for i in range(2,max_raw+1):
        dict1 = dict(
            case_id = sheet.cell(row=i,column=1).value,
            method=sheet.cell(row=i, column=4).value,
            url = sheet.cell(row=i,column=5).value,
            data = sheet.cell(row=i,column=6).value,
            path=sheet.cell(row=i, column=7).value,
            expect = sheet.cell(row=i,column=8).value
        )
        case_list.append(dict1)
    return case_list

def responseToJson(response):
    return json.loads(response.text)


def testdata():
    #fake = Factory.create('zh_CN')
    fake = Faker('zh_CN')

    name_list = []
    for i in range(0, 10):
        name_list.append(fake.name())
        name_list.append(fake.address())
        name_list.append(fake.text())

    return name_list

'''
class CustomProvider(BaseProvider):
    def product_name(self):
        products = ["Widget", "Gadget", "Doodad", "Thingamajig"]
        return self.random_element(products)

        # 添加自定义提供者到 Faker

fake.add_provider(CustomProvider)

# 使用自定义提供者生成虚拟产品名称
product = fake.product_name()
print("虚拟产品名称:", product)


class CustomFaker(Faker):
    def custom_method(self):
        return "Custom Data"

# 使用自定义虚假数据生成器
custom_fake = CustomFaker()

# 生成自定义虚假数据
custom_data = custom_fake.custom_method()

print("自定义虚假数据:", custom_data)

class CustomProvider(BaseProvider):
    def product_name(self):
        products = ["Widget", "Gadget", "Doodad", "Thingamajig"]
        return self.random_element(products)

# 添加自定义提供者到 Faker
fake.add_provider(CustomProvider)

# 使用自定义提供者生成虚拟产品名称
product = fake.product_name()
print("虚拟产品名称:", product)

class CustomFaker(Faker):
    def custom_method(self):
        return "Custom Data"

# 使用自定义虚假数据生成器
custom_fake = CustomFaker()
# 生成自定义虚假数据
custom_data = custom_fake.custom_method()
print("自定义虚假数据:", custom_data)

'''